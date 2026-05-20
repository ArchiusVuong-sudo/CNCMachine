#!/usr/bin/env python
"""Dump .xls/.xlsx cells with column letters preserved (handles old BIFF .xls + .xlsx).
Usage: python xls_dump.py "<file>" [maxrows_per_sheet]
Setup-sheet tool tables use columns A..P (see reference/data_dictionary.md)."""
import sys, warnings
warnings.filterwarnings("ignore")
from pathlib import Path

def col(i):  # 0-based -> A, B, ... Z, AA
    s = ""
    i += 1
    while i:
        i, r = divmod(i - 1, 26)
        s = chr(65 + r) + s
    return s

def rows_xlsx(p):
    import openpyxl
    wb = openpyxl.load_workbook(p, data_only=True)
    for sn in wb.sheetnames:
        yield sn, [list(r) for r in wb[sn].iter_rows(values_only=True)]
    wb.close()

def rows_xls(p):
    import xlrd
    wb = xlrd.open_workbook(p)
    for sn in wb.sheet_names():
        ws = wb.sheet_by_name(sn)
        yield sn, [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]

def main():
    p = Path(sys.argv[1])
    maxr = int(sys.argv[2]) if len(sys.argv) > 2 else 120
    gen = rows_xlsx(p) if p.suffix.lower() == ".xlsx" else rows_xls(p)
    for sn, rows in gen:
        ne = [i for i, r in enumerate(rows) if any(str(c).strip() for c in r if c is not None)]
        print(f"\n##### sheet '{sn}' ({len(rows)} rows, {len(ne)} non-empty)")
        shown = 0
        for ri in ne:
            cells = ["" if c is None else str(c).strip() for c in rows[ri]]
            disp = " | ".join(f"{col(i)}={c}" for i, c in enumerate(cells) if c != "")
            if disp:
                print(f"R{ri}: {disp[:320]}")
                shown += 1
            if shown >= maxr:
                print("  ...(truncated)")
                break

if __name__ == "__main__":
    main()
